import pandas as pd
from canvasapi import Canvas
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import urllib3
import re

# Suppress the Mac LibreSSL warning
urllib3.disable_warnings(urllib3.exceptions.NotOpenSSLWarning)

# ==================== CONFIGURATION ====================
API_URL = "https://vsu.instructure.com"
API_KEY = "" # Use a NEW key here!
# =======================================================

def get_category(name):
    if not name: return 'Assignment'
    n = str(name).lower()
    if any(x in n for x in ['exam', 'test', 'midterm', 'final']): return 'Exam/Test'
    if 'quiz' in n: return 'Quiz'
    if any(x in n for x in ['lab', 'experiment', 'workshop']): return 'Lab'
    if any(x in n for x in ['project', 'paper', 'essay', 'portfolio']): return 'Project'
    if any(x in n for x in ['homework', 'hw', 'reading', 'problem']): return 'Homework'
    if any(x in n for x in ['participation', 'attendance', 'discussion']): return 'In-Class/Disc'
    return 'Assignment'

def export_to_excel():
    try:
        # Initialize Canvas
        canvas = Canvas(API_URL, API_KEY)
        user = canvas.get_current_user()
        courses = user.get_courses(enrollment_state='active')
        
        data_rows = []
        now = datetime.now(timezone.utc) # Timezone-aware "Now"

        print(f"--- Fetching Data for: {user.name} ---")

        for course in courses:
            c_name = getattr(course, 'name', 'Unknown Course')
            print(f"Checking: {c_name}...")
            
            try:
                assignments = list(course.get_assignments())
                if not assignments:
                    data_rows.append({
                        'Status': "⚪️ Empty", 'Course': c_name, 'Type': "N/A",
                        'Assignment': "No assignments found", 'Days Left': None,
                        'Due Date': None, 'Points': 0, 'Description': "No content.", 'Link': ""
                    })
                    continue

                for a in assignments:
                    due_date_raw = getattr(a, 'due_at', None)
                    days_left = None
                    
                    if due_date_raw:
                        # Convert to timezone-aware datetime for safe comparison
                        due_date_obj = pd.to_datetime(due_date_raw).tz_localize(None).tz_localize(timezone.utc)
                        days_left = (due_date_obj - now).days

                    raw_desc = getattr(a, 'description', '')
                    clean_desc = re.sub('<[^<]+?>', '', str(raw_desc)) if raw_desc else "No description."
                    
                    data_rows.append({
                        'Status': "✅ Done" if getattr(a, 'has_submitted_submissions', False) else "⏳ Pending",
                        'Course': c_name, 'Type': get_category(getattr(a, 'name', '')),
                        'Assignment': getattr(a, 'name', 'Unnamed'), 'Days Left': days_left,
                        'Due Date': due_date_raw, 'Points': getattr(a, 'points_possible', 0),
                        'Description': clean_desc[:1000].strip(), 'Link': getattr(a, 'html_url', '')
                    })
            except Exception: continue

        df = pd.DataFrame(data_rows)
        if df.empty:
            print("No assignments found.")
            return

        # Safe Date Conversion
        df['Due Date DT'] = pd.to_datetime(df['Due Date'], errors='coerce')
        # Ensure the date column is timezone-aware to match 'now'
        if not df['Due Date DT'].dropna().empty:
            df['Due Date DT'] = df['Due Date DT'].dt.tz_localize(None).dt.tz_localize(timezone.utc)

        # Formatting for display
        df['Due Date Display'] = df['Due Date DT'].dt.strftime('%b %d, %I:%M %p').fillna('—')
        
        # Sort logic
        df['sort_days'] = pd.to_numeric(df['Days Left'], errors='coerce').fillna(999)
        df = df.sort_values(by=['Status', 'sort_days'], ascending=[False, True])

        # Prepare final columns
        final_df = df.drop(columns=['Due Date', 'Due Date DT', 'sort_days']).rename(columns={'Due Date Display': 'Due Date'})
        
        filename = "Canvas_Planner.xlsx"
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        final_df.to_excel(writer, index=False, sheet_name='Dashboard')
        
        # --- STYLING ---
        ws = writer.sheets['Dashboard']
        header_color, row_alt_color = "334155", "F8FAFC"
        urgent_red, safe_green, border_color = "FEE2E2", "DCFCE7", "E2E8F0"
        
        header_font = Font(name='Helvetica Neue', size=11, bold=True, color="FFFFFF")
        main_font = Font(name='Helvetica Neue', size=10, color="1E293B")
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        alt_fill = PatternFill(start_color=row_alt_color, end_color=row_alt_color, fill_type="solid")
        red_fill = PatternFill(start_color=urgent_red, end_color=urgent_red, fill_type="solid")
        green_fill = PatternFill(start_color=safe_green, end_color=safe_green, fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color=border_color),
                            right=Side(style='thin', color=border_color),
                            top=Side(style='thin', color=border_color),
                            bottom=Side(style='thin', color=border_color))

        for col in range(1, len(final_df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill, cell.font, cell.border = header_fill, header_font, thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in range(2, len(final_df) + 2):
            is_alt = row % 2 == 0
            for col in range(1, len(final_df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font, cell.border = main_font, thin_border
                cell.alignment = Alignment(vertical='top', horizontal='left')
                if is_alt: cell.fill = alt_fill
            
            days_cell = ws.cell(row=row, column=5)
            try:
                if days_cell.value is not None:
                    val = int(days_cell.value)
                    if val < 3: days_cell.fill = red_fill
                    elif val > 7: days_cell.fill = green_fill
            except: pass

            ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 35

        widths = [10, 25, 15, 35, 12, 22, 10, 60, 30]
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w

        writer.close()
        print(f"\n✨ SUCCESS: '{filename}' has been created.")

    except Exception as e:
        print(f"\n❌ ERROR: {e}")

if __name__ == "__main__":
    export_to_excel()